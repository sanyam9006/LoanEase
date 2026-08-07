"""Lead management API — create, list, export, get single lead."""

import io
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import or_, func
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.lead import Lead
from app.models.user import User
from app.schemas.lead import LeadCreate, LeadResponse, LeadDetail, LeadListResponse
from app.services.credit_score import fetch_credit_score
from app.services.bre_engine import evaluate_lead

router = APIRouter(prefix="/api/leads", tags=["Leads"])


@router.post("", response_model=LeadResponse)
def create_lead(lead_data: LeadCreate, db: Session = Depends(get_db)):
    """
    POST /api/leads
    Main pipeline: validate -> check duplicate -> fetch credit score -> run BRE -> save lead
    """
    # check if mobile already exists
    existing = db.query(Lead).filter(Lead.mobile == lead_data.mobile).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Lead already exists",
        )

    # get credit score
    credit_result = fetch_credit_score(lead_data.full_name, lead_data.mobile)

    credit_score = credit_result.score if credit_result.success else None
    credit_status = "success" if credit_result.success else "failed"

    # run BRE rules
    bre_data = {
        "dob": lead_data.dob,
        "monthly_income": lead_data.monthly_income,
        "credit_score": credit_score,
        "loan_amount": lead_data.loan_amount,
        "property_value": lead_data.property_value,
    }
    bre_result = evaluate_lead(db, bre_data)

    # save to db
    new_lead = Lead(
        full_name=lead_data.full_name,
        mobile=lead_data.mobile,
        email=lead_data.email,
        dob=lead_data.dob,
        city=lead_data.city,
        pincode=lead_data.pincode,
        loan_type=lead_data.loan_type,
        employment_type=lead_data.employment_type,
        monthly_income=lead_data.monthly_income,
        loan_amount=lead_data.loan_amount,
        property_value=lead_data.property_value,
        credit_score=credit_score,
        credit_score_status=credit_status,
        bre_status=bre_result.status,
        rejection_reasons=bre_result.rejection_reasons if bre_result.rejection_reasons else None,
        consent=lead_data.consent,
    )
    db.add(new_lead)
    db.commit()
    db.refresh(new_lead)

    return LeadResponse(
        status="success",
        lead_id=new_lead.id,
        credit_score=credit_score,
        bre_status=bre_result.status,
        rejection_reasons=bre_result.rejection_reasons if bre_result.rejection_reasons else None,
        message=credit_result.error if not credit_result.success else None,
    )


@router.get("", response_model=LeadListResponse)
def list_leads(
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=100),
    search: Optional[str] = Query(None),
    loan_type: Optional[str] = Query(None),
    bre_status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """GET /api/leads — paginated list with search & filters. Admin only."""
    query = db.query(Lead)

    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                Lead.full_name.ilike(search_term),
                Lead.mobile.ilike(search_term),
                Lead.email.ilike(search_term),
            )
        )

    if loan_type:
        query = query.filter(Lead.loan_type == loan_type)

    if bre_status:
        query = query.filter(Lead.bre_status == bre_status)

    total = query.count()

    leads = (
        query.order_by(Lead.created_at.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    lead_details = []
    for lead in leads:
        lead_details.append(
            LeadDetail(
                id=lead.id,
                full_name=lead.full_name,
                mobile=lead.mobile,
                email=lead.email,
                dob=lead.dob,
                city=lead.city,
                pincode=lead.pincode,
                loan_type=lead.loan_type,
                employment_type=lead.employment_type,
                monthly_income=lead.monthly_income,
                loan_amount=lead.loan_amount,
                property_value=lead.property_value,
                credit_score=lead.credit_score,
                credit_score_status=lead.credit_score_status,
                bre_status=lead.bre_status,
                rejection_reasons=lead.rejection_reasons,
                consent=lead.consent,
                created_at=(
                    lead.created_at.strftime("%Y-%m-%dT%H:%M:%SZ")
                    if lead.created_at
                    else ""
                ),
            )
        )

    total_pages = (total + per_page - 1) // per_page

    return LeadListResponse(
        leads=lead_details,
        total=total,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
    )


@router.get("/export")
def export_leads_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export all leads as an Excel file. Admin only."""
    import openpyxl

    leads = db.query(Lead).order_by(Lead.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = [
        "Lead ID", "Full Name", "Mobile", "Email", "DOB", "City", "Pincode",
        "Loan Type", "Employment Type", "Monthly Income", "Loan Amount",
        "Property Value", "Credit Score", "BRE Status", "Rejection Reasons",
        "Created Date",
    ]
    ws.append(headers)

    # style the header row
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for lead in leads:
        reasons = ", ".join(lead.rejection_reasons) if lead.rejection_reasons else ""
        ws.append([
            lead.id, lead.full_name, lead.mobile, lead.email, lead.dob,
            lead.city, lead.pincode, lead.loan_type, lead.employment_type,
            lead.monthly_income, lead.loan_amount, lead.property_value,
            lead.credit_score, lead.bre_status, reasons,
            lead.created_at.strftime("%Y-%m-%d %H:%M") if lead.created_at else "",
        ])

    # auto-fit columns
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=leads_export.xlsx"},
    )


@router.get("/{lead_id}", response_model=LeadDetail)
def get_lead(
    lead_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get a single lead by ID. Admin only."""
    lead = db.query(Lead).filter(Lead.id == lead_id).first()
    if not lead:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Lead not found",
        )

    return LeadDetail(
        id=lead.id,
        full_name=lead.full_name,
        mobile=lead.mobile,
        email=lead.email,
        dob=lead.dob,
        city=lead.city,
        pincode=lead.pincode,
        loan_type=lead.loan_type,
        employment_type=lead.employment_type,
        monthly_income=lead.monthly_income,
        loan_amount=lead.loan_amount,
        property_value=lead.property_value,
        credit_score=lead.credit_score,
        credit_score_status=lead.credit_score_status,
        bre_status=lead.bre_status,
        rejection_reasons=lead.rejection_reasons,
        consent=lead.consent,
        created_at=(
            lead.created_at.strftime("%Y-%m-%dT%H:%M:%SZ")
            if lead.created_at
            else ""
        ),
    )
